from io import BytesIO 
import os
import requests

from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, logout
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, HttpResponse
from django.views.generic import ListView
from rest_framework import generics, status
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
# from rest_framework_simplejwt.authentication import JWTAuthentication
from PIL import Image as PILImage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, TABLOID
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from .forms import CoursesForm, UpdateCourseForm, viewCourseForm
from .models import AdminLogin, Course, PartnerLogo, Testimonial, PlacementStory, FAQ, Blog, Career
from .serializers import CourseSerializer, UserSerializer

from django.http import HttpResponse
from django.template import loader
from django.utils.html import strip_tags
# from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
# from rest_framework_simplejwt.serializers import TokenObtainPairSerializer, TokenRefreshSerializer

from rest_framework.exceptions import AuthenticationFailed
import jwt, datetime
from .authentication import JWTAuthentication
from rest_framework.decorators import api_view, permission_classes
# ----------------------------- Admin Views ----------------------------



def admin_login_view(request):
    return render(request, 'Admin_Login_App/AdminLogin.html')

def admin_login_submit(request):
    # print(request)
    return HttpResponse("Welcome")
    # return render(request, 'Admin_Login_App/AdminLogin.html')

def dashboard_view(request):
    return render(request, 'Admin_Login_App/Admin_Body.html')

def logout_view(request):
    if request.user.is_authenticated:
        logout(request)
    return redirect('admin_login')

def user_logout(request):
    if request.user.is_authenticated:
        logout(request)
    return redirect('home')

# ---------------------------- Admin Login API ----------------------------

# Create your views here.
class RegisterView(APIView):
    def post(self, request):
        serializer = UserSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

class AdminLoginAPI(APIView):
    
    # def post(self, request):
    #     username = request.data.get('username')
    #     print(username)
    #     password = request.data.get('password')
    #     print(password)
    #     user = authenticate(username=username, password=password)
    #     if user:
    #         refresh = RefreshToken.for_user(user)
    #         return Response({
    #             'access': str(refresh.access_token),
    #             'refresh': str(refresh)
    #         })
    #     return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

    def post(self, request):
        username = request.data['username']
        password = request.data['password']
        
        print(username)
        print(password)

        # user = AdminLogin.objects.filter(email=email).first()
        user = AdminLogin.objects.filter(username = username).first()

        if user is None:
            raise AuthenticationFailed('User not found!')

        if not user.check_password(password):
            raise AuthenticationFailed('Incorrect password!')

        payload = {
            'id': user.id,
            'exp': datetime.datetime.utcnow() + datetime.timedelta(minutes=120),
            'iat': datetime.datetime.utcnow()
        }
    

        token = jwt.encode(payload, 'secret', algorithm='HS256')
    
        # token = jwt.encode(payload, 'secret', algorithm='HS256').decode('utf-8')
        # token = jwt.decode(jwt.encode(payload, 'secret', algorithm='HS256'), 'secret', algorithms=['HS256'])

        response = Response()

        response.set_cookie(key='jwt', value=token, httponly=True, max_age=24 * 60 * 60)
        response.data = {
            'jwt': token
        }
        return response

# ---------------------------- Admin User Info API ----------------------------

class get_admin_usernames(APIView):
    # permission_classes = [IsAuthenticated]

    # def get(self, request):
    #     admin_usernames = AdminLogin.objects.values_list('username', flat=True)
    #     return Response(list(admin_usernames))

    def get(self, request):
        token = request.COOKIES.get('jwt')

        if not token:
            raise AuthenticationFailed('Unauthenticated!')

        try:
            payload = jwt.decode(jwt=token, key='secret', algorithms=['HS256'])

        except jwt.ExpiredSignatureError:
            raise AuthenticationFailed('Unauthenticated!')

        user = AdminLogin.objects.filter(id=payload['id']).first()
        serializer = UserSerializer(user)
        return Response(serializer.data)

class LogoutView(APIView):
    def post(self, request):
        response = Response()
        response.delete_cookie('jwt')
        response.data = {
            'message': 'success'
        }
        return response

# ---------------------------- Course Views ---------------------------------

class KeywordListView(ListView):
    paginate_by = 6
    model = Course

# def course_view(request):
#     # Get all courses ordered by title
#     data = Course.objects.order_by('Title').all()
    
#     # Get the current page number from the request
#     page = request.GET.get('page', 1)
    
#     # Create a Paginator object with 3 items per page
#     paginator = Paginator(data, per_page=3)
    
#     # Get the page object for the current page
#     page_object = paginator.get_page(page)
    
#     # Adjust the elided pages range
#     page_object.adjusted_elided_pages = paginator.get_elided_page_range(page)
    
#     # Prepare the context for rendering the template
#     context = {"page_obj": page_object}
    
#     # Render the template with the context
#     return render(request, 'Admin_Login_App/courses.html', context)


def course_view(request, page=1, per_page=1):
    # Get all courses ordered by title
    data = Course.objects.order_by('id').all()
    
    # Create a Paginator object with 3 items per page
    paginator = Paginator(data, per_page)
    
    # Get the page object for the requested page
    try:
        page_object = paginator.page(page)
    except PageNotAnInteger:
        page_object = paginator.page(1)
    except EmptyPage:
        page_object = paginator.page(paginator.num_pages)
    
    # Prepare the context for rendering the template
    context = {
        "page_obj": page_object
    }
    
    # Render the template with the context
    return render(request, 'Admin_Login_App/courses.html', context)


# def listing(request, page):
#     data = Course.objects.order_by('Title').all()
    
#     paginator = Paginator(data, per_page=3)
    
#     page_object = paginator.get_page(page)
#     page_object.adjusted_elided_pages = paginator.get_elided_page_range(page)
#     context = {"page_obj": page_object}
#     return render(request, 'Admin_Login_App/courses.html', context)

# def listing(request, page, per_page=3):
#     data = Course.objects.all().order_by('id')
    
#     paginator = Paginator(data, per_page)
    
#     try:
#         page_object = paginator.page(page)
#     except PageNotAnInteger:
#         # If page is not an integer, deliver first page.
#         page_object = paginator.page(1)
#     except EmptyPage:
#         # If page is out of range (e.g. 9999), deliver last page of results.
#         page_object = paginator.page(paginator.num_pages)
    
#     # Optionally adjust for elided pages if needed
#     page_object.adjusted_elided_pages = paginator.get_elided_page_range(page)
    
#     context = {
#         "page_obj": page_object,
#         "per_page": per_page,  # Pass per_page to the context
#         "paginator": paginator,
#     }
#     return render(request, 'Admin_Login_App/courses.html', context)

# def listing(request):
#     courses = Course.objects.all().order_by('id')  # Ordering by 'id' as an example
#     paginator = Paginator(courses, per_page=3)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#     return render(request, 'Admin_Login_App/courses.html', {'page_obj': page_obj})

def course_page_view(request):
    form = CoursesForm()
    return render(request, 'Admin_Login_App/course_page.html', {'form': form})

def navbar_save_view(request):
    return render(request, 'Admin_Login_App/navbar_save_course.html')

def navbar_view_course(request):
    return render(request, 'Admin_Login_App/navbar_view_course.html')

def update_course(request, id):
    course = Course.objects.get(id=id)
    form = UpdateCourseForm(instance=course)
    return render(request, 'Admin_Login_App/course_update.html', {'form': form, 'datas': course})

def view_course(request, id):
    course = get_object_or_404(Course, id=id)
    return render(request, 'Admin_Login_App/view_course.html', {'datas': course})

def delete_course(request, id):
    obj = Course.objects.get(id=id)
    return render(request, 'Admin_Login_App/navbar.html', {'datas': obj})

# ---------------------------- Course API -----------------------------------

class AllKeywordsView(ListView):
    model = Course
    template_name = "Admin_Login_App/course_list.html"

# def listing_api(request):
#     page_number = int(request.GET.get("page", 1))
#     per_page = int(request.GET.get("per_page", 3))
#     startswith = request.GET.get("startswith", "")
#     courses = Course.objects.filter(
#         Title__startswith=startswith
#     )
#     paginator = Paginator(courses, per_page)
#     page_obj = paginator.get_page(page_number)
#     data = [{"id": kw.id, "Title": kw.Title, "Technologies": kw.Technologies, "Description": kw.Description ,"Images": kw.Images.url, "status": kw.status} for kw in page_obj.object_list]

#     pagination = {
#         "current": page_obj.number,
#         "has_next": page_obj.has_next(),
#         "has_previous": page_obj.has_previous(),
#         "total_pages": paginator.num_pages,
#         "page_range": list(paginator.get_elided_page_range(page_number)),
#     }

#     payload = {
#         "data": data,
#         "pagination": pagination
#     }
#     return JsonResponse(payload)


def listing_api(request):
    page_number = request.GET.get("page", 1)
    per_page = request.GET.get("per_page", 1)
    startswith = request.GET.get("startswith", "")
    
    # Filtering courses based on startswith
    courses = Course.objects.filter(id__startswith=startswith)
    total_items = courses.count()
    # Create a Paginator object with per_page items per page
    paginator = Paginator(courses, per_page)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Define functions to clean description and technologies
    def clean_description(description):
        # Remove HTML tags using regex
        clean_description = re.sub(r'<[^>]*>', '', description)
        return clean_description.strip()  # Strip any leading or trailing whitespace
    
    def clean_technologies(technologies):
        # Remove unwanted commas using regex
        clean_technologies = re.sub(r',+$', '', technologies)
        return clean_technologies.strip()  # Strip any leading or trailing whitespace
    
    # Generate JSON response with cleaned fields
    data = [{
        "id": course.id,
        "Title": course.Title,
        "Technologies": clean_technologies(course.Technologies),
        "Description": clean_description(course.Description),
        "Images": course.Images.url,  # Assuming Images field is a FileField or ImageField
        "status": course.status
    } for course in page_obj.object_list]

    pagination = {
        "current": page_obj.number,
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
        "total_pages": paginator.num_pages,
        "page_range": list(paginator.get_elided_page_range(page_number)),
    }

    payload = {
        "data": data,
        "pagination": pagination,
        "total_items": total_items,
    }
    
    # Return JSON response
    return JsonResponse(payload)

class CourseListCreateView(generics.ListCreateAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

class CourseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class CourseUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    partial = True
    
    
class CourseDeleteView(generics.DestroyAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response(print("delete Movie"))

# ----------------------------- PDF Generation ----------------------------

# def pdf(request):
#     data = Course.objects.all()
    
#     buffer = BytesIO()
#     doc = SimpleDocTemplate(buffer, pagesize=TABLOID)
#     elements = []

#     # Define the table data
#     table_data = [["Course Name", "Technologies", "Description", "Image"]]
    
#     for course_obj in data:
#         # Retrieve the image
#         image_path = ""
#         image_data = None
#         print(course_obj.Images.url)
#         try:
#             response = requests.get(course_obj.Images.url, stream=True)
#             if response.status_code == 200:
#                 image = PILImage.open(response.raw)
#                 image_path = f"temp_image_{course_obj.id}.png"
#                 image.save(image_path)
#                 image_data = Image(image_path, 2*inch, 2*inch)
#         except Exception as e:
#             image_data = Paragraph("Image not available", getSampleStyleSheet()['Normal'])

#         # Create a Paragraph for the description to handle overflow
#         description = Paragraph(course_obj.Description, getSampleStyleSheet()['Normal'])

#         # Add the course data to the table
#         course_row = [
#             course_obj.Title,
#             course_obj.Technologies,
#             description,
#             image_data
#         ]
#         table_data.append(course_row)

#     # Create the table
#     table = Table(table_data, colWidths=[2*inch, 1.5*inch, 3.5*inch, 2*inch])

#     # Add a table style
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#         ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('FONTSIZE', (0, 0), (-1, 0), 14),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#         ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#     ]))

#     elements.append(table)
#     doc.build(elements)

#     buffer.seek(0)
#     x = datetime.datetime.now()
#     date_format = x.strftime("%Y-%m-%d")
#     time_format = x.strftime("%H-%M-%S")
#     save_path = f'pdf/course_list_{date_format}_{time_format}.pdf'
    
#     with open(save_path, 'wb') as f:
#         f.write(buffer.getvalue())
    
#     return HttpResponse(f"PDF file has been generated and saved at: {save_path}")


from django.http import Http404

from . import renderers

def clean_html(raw_html):
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext

# def pdf(request, *args, **kwargs):
#     data = Course.objects.all()
    
#     if not data:
#         raise Http404("No courses available.")
    
#     content_list = []
#     for course_obj in data:
#         cleaned_description = clean_html(course_obj.Description).replace(',', '')
#         cleaned_technologies = clean_html(course_obj.Technologies).replace(',', '')
        
#         content_list.append({
#             'Course_Name': course_obj.Title, 
#             'Technologies': cleaned_technologies,
#             'Description': cleaned_description,
#             'Images': course_obj.Images,
#         })
#         print(course_obj.Images)
    
#     content = {'courses': content_list}
#     return renderers.render_to_pdf('Admin_Login_App/course_data_list.html', content)

import csv
from django.template import loader
from django.http import HttpResponse

import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, Border, Side
from PIL import Image as PILImage
from io import BytesIO
import re

def clean_text(text):
    """Remove unwanted commas and HTML tags from the text."""
    # Remove HTML tags
    clean = re.compile('<.*?>')
    text = re.sub(clean, '', text)
    
    # Remove unwanted commas (you can define what "unwanted" means in your context)
    text = text.replace(',', '')

    return text

# def excel_view(request):
#     data = Course.objects.all()

#     # Create an Excel workbook
#     wb = openpyxl.Workbook()
#     ws = wb.active

#     # Define header row with font style and alignment
#     header_row = ['Course Name', 'Topics', 'Description', 'Images']
#     ws.append(header_row)
#     for cell in ws[1]:
#         cell.font = Font(bold=True, color='000000')
#         cell.alignment = Alignment(horizontal='left', vertical='center')

#     # Set column widths for better readability
#     column_widths = [20, 30, 50, 20]
#     for i, width in enumerate(column_widths, start=1):
#         ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

#     # Add data rows with alignment and borders
#     thin_border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )

#     for idx, item in enumerate(data, start=2):
#         cleaned_topics = clean_text(item.Technologies)
#         cleaned_description = clean_text(item.Description)
#         ws.append([item.Title, cleaned_topics, cleaned_description])

#         # Align text and apply borders
#         for cell in ws[idx]:
#             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
#             cell.border = thin_border

#         # Add image to the worksheet if it exists
#         if item.Images:
#             image_path = item.Images.path  # Get the file path of the image

#             # Resize the image and add padding
#             with PILImage.open(image_path) as img:
#                 max_width = 50  # New max width after considering padding
#                 max_height = 50  # New max height after considering padding
#                 img.thumbnail((max_width, max_height))

#                 # Add padding
#                 padding = 10  # 10px padding on each side
#                 padded_img = PILImage.new("RGBA", (img.width + 2 * padding, img.height + 2 * padding), (255, 255, 255, 0))
#                 padded_img.paste(img, (padding, padding))

#                 # Save the padded image to a BytesIO object
#                 image_stream = BytesIO()
#                 padded_img.save(image_stream, format='PNG')
#                 image_stream.seek(0)

#                 # Create an ExcelImage object from the BytesIO object
#                 excel_img = ExcelImage(image_stream)

#                 # Adjust the row height to match the image height
#                 row_height = padded_img.height * 0.75  # Adjust the scaling factor if needed
#                 ws.row_dimensions[idx].height = row_height

#                 # Center the image in the cell
#                 col_letter = 'D'
#                 col_width = ws.column_dimensions[col_letter].width
#                 x_offset = (col_width * 7.5 - excel_img.width) / 2  # Approx 7.5 pixels per Excel column unit
#                 y_offset = (row_height - excel_img.height) / 2

#                 # Anchor the image to the cell
#                 excel_img.anchor = f'{col_letter}{idx}'
#                 ws.add_image(excel_img)

#     # Create an in-memory file-like object to save the workbook
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     # Create the HTTP response with Excel content type and attachment header
#     response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
    
#     return response

def clean_html(raw_html):
    cleanr = re.compile('<.*?>')
    cleantext = re.sub(cleanr, '', raw_html)
    return cleantext


def clean_html(html):
    # Function to clean HTML content by stripping tags
    return strip_tags(html)

import csv
from django.http import HttpResponse

# def csv_view(request):
#     if request.method == 'POST':
#         ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

#         # Create the HttpResponse object with the appropriate CSV header
#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="course_list_csv.csv"'

#         writer = csv.writer(response)
        
#         # Write the header row
#         writer.writerow(['Course Name', 'Technologies', 'Description', 'Image URL'])

#         # Assuming you have a Course model, adjust this part accordingly
#         # from .models import Course

#         # Fetch selected courses based on IDs
#         selected_courses = Course.objects.filter(id__in=ids)
        
#         for course in selected_courses:
#             # Clean and prepare the data for each row
#             cleaned_description = clean_html(course.Description).replace(',', '')
#             cleaned_technologies = clean_html(course.Technologies).replace(',', '')
#             img_path = course.Images
#             image_url = f"http://127.0.0.1:8000/media/{img_path}"

#             # Write the data row
#             writer.writerow([course.Title, cleaned_technologies, cleaned_description, image_url])

#         return response

#     # Handle GET request or non-AJAX POST request here if needed
#     return HttpResponse(status=400)  # Bad request if not POST or AJAX



# def csv_view(request):
#     if request.method == 'POST':
#         ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

#         # Create the HttpResponse object with the appropriate CSV header
#         response = HttpResponse(content_type='text/csv')
#         response['Content-Disposition'] = 'attachment; filename="course_list_csv.csv"'

#         writer = csv.writer(response)
        
#         # Write the header row
#         writer.writerow(['Course Name', 'Technologies', 'Description', 'Image URL'])

#         # Fetch selected courses based on IDs
#         selected_courses = Course.objects.filter(id__in=ids)
        
#         for course in selected_courses:
#             # Clean and prepare the data for each row
#             cleaned_description = clean_html(course.Description).replace(',', '')  # Replace with actual field names
#             cleaned_technologies = clean_html(course.Technologies).replace(',', '')  # Replace with actual field names
#             img_path = course.Images.url if course.Images else ''  # Replace with actual field names and image URL logic
#             image_url = request.build_absolute_uri(img_path)

#             # Write the data row
#             writer.writerow([course.Title, cleaned_technologies, cleaned_description, image_url])

#         return response

#     # Handle GET request or non-AJAX POST request here if needed
#     return HttpResponse(status=400)  # Bad request if not POST or AJAX


import csv
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def export_courses_csv(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="course_list_csv.csv"'

        writer = csv.writer(response)

        # Write the header row
        writer.writerow(['Course Name', 'Technologies', 'Description', 'Image URL'])

        # Fetch selected courses based on IDs
        selected_courses = Course.objects.filter(id__in=ids)

        for course in selected_courses:
            # Clean and prepare the data for each row
            cleaned_description = clean_html(course.Description).replace(',', '')  # Replace with actual field names
            cleaned_technologies = clean_html(course.Technologies).replace(',', '')  # Replace with actual field names
            img_path = course.Images.url if course.Images else ''  # Replace with actual field names and image URL logic
            image_url = request.build_absolute_uri(img_path)

            # Write the data row
            writer.writerow([course.Title, cleaned_technologies, cleaned_description, image_url])

        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX


import openpyxl
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def export_courses_excel(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Fetch selected courses based on IDs
        selected_courses = Course.objects.filter(id__in=ids)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define header row with font style and alignment
        header_row = ['Course Name', 'Topics', 'Description', 'Images']
        ws.append(header_row)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='000000')
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # Set column widths for better readability
        column_widths = [20, 30, 50, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Add data rows with alignment and borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for idx, item in enumerate(selected_courses, start=2):
            cleaned_topics = clean_text(item.Technologies)
            cleaned_description = clean_text(item.Description)
            ws.append([item.Title, cleaned_topics, cleaned_description])

            # Align text and apply borders
            for cell in ws[idx]:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.border = thin_border

            # Add image to the worksheet if it exists
            if item.Images:
                image_path = item.Images.path  # Get the file path of the image

                # Resize the image and add padding
                with PILImage.open(image_path) as img:
                    max_width = 50  # New max width after considering padding
                    max_height = 50  # New max height after considering padding
                    img.thumbnail((max_width, max_height))

                    # Add padding
                    padding = 10  # 10px padding on each side
                    padded_img = PILImage.new("RGBA", (img.width + 2 * padding, img.height + 2 * padding), (255, 255, 255, 0))
                    padded_img.paste(img, (padding, padding))

                    # Save the padded image to a BytesIO object
                    image_stream = BytesIO()
                    padded_img.save(image_stream, format='PNG')
                    image_stream.seek(0)

                    # Create an ExcelImage object from the BytesIO object
                    excel_img = ExcelImage(image_stream)

                    # Adjust the row height to match the image height
                    row_height = padded_img.height * 0.75  # Adjust the scaling factor if needed
                    ws.row_dimensions[idx].height = row_height

                    # Center the image in the cell
                    col_letter = 'D'
                    col_width = ws.column_dimensions[col_letter].width
                    x_offset = (col_width * 7.5 - excel_img.width) / 2  # Approx 7.5 pixels per Excel column unit
                    y_offset = (row_height - excel_img.height) / 2

                    # Anchor the image to the cell
                    excel_img.anchor = f'{col_letter}{idx}'
                    ws.add_image(excel_img)

        # Create an in-memory file-like object to save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response with Excel content type and attachment header
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
        
        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def export_courses_pdf(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # # Fetch selected courses based on IDs
        selected_courses = Course.objects.filter(id__in=ids)
        
        if not selected_courses:
            raise Http404("No courses available.")
        
        content_list = []
        for course_obj in selected_courses:
            cleaned_description = clean_html(course_obj.Description).replace(',', '')
            cleaned_technologies = clean_html(course_obj.Technologies).replace(',', '')
            
            content_list.append({
                'Course_Name': course_obj.Title, 
                'Technologies': cleaned_technologies,
                'Description': cleaned_description,
                'Images': course_obj.Images,
            })
            print(course_obj.Images)
        
        content = {'courses': content_list}
        return renderers.render_to_pdf('Admin_Login_App/course_data_list.html', content)

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX
